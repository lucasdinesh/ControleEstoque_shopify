import shopify
import requests
import simplejson
import openpyxl
from openpyxl.styles import  PatternFill

class ShopifyApp:
    def __init__(self):
       # self.chaveapi =
        #self.senhaapi =
        self.apiversion = '2020-07'
        self.proxpage=''
       # self.shop_url = f"https://{self.chaveapi}:{self.senhaapi}@.myshopify.com/admin/api/{self.apiversion}"
        shopify.ShopifyResource.set_site(self.shop_url)

    def get_pedidos(self, page):
        try:
            if page == 1:
                url_pedidos = f"{self.shop_url}/orders.json?limit=100&status=any"

                url_resposta = requests.get(url_pedidos)

                url_header = requests.head(url=url_pedidos)

                self.proxpage = url_header.links["next"]['url']

                orders = simplejson.loads(url_resposta.text)

                return orders
            else:
                url_pedidos = f"{self.shop_url}{self.proxpage[47:]}"

                url_resposta = requests.get(url_pedidos)

                url_header = requests.head(url=url_pedidos)

                self.proxpage = url_header.links['next']['url']

                orders = simplejson.loads(url_resposta.text)

                return orders
        except Exception as error:
            print('erro ao executar pedidos', error)

    def valida_endereco(self, end1, end2):
        if (end1.isalpha() and end2.isalpha()) == True:
            return False
        else:
            return True

    def cancelado(self, pedido, planilha1):
        cancelado_tupla = (pedido, 'CANCELADO', 'CANCELADO', 'CANCELADO', 'CANCELADO', 'CANCELADO', 'CANCELADO')
        planilha1.append(cancelado_tupla)

    def insere_tupla(self, pedido, status, forma_envio, quantidade_itens, forma_pagamento, processado,
                     planilha1=None, planilha2=None):
        insere_tupla = (pedido, status, forma_envio, quantidade_itens, forma_pagamento, processado, " ")
        if planilha1 != None:
            planilha1.append(insere_tupla)
        else:
            planilha2.append(insere_tupla)

    def colore_celula(self, planilha):
        iterador = 0
        greenFill = PatternFill(patternType="solid", start_color="00FF00")
        redFill = PatternFill(patternType='solid', start_color='FF0000')
        yellowFill = PatternFill(patternType='solid', start_color='FFFF00')
        planilha = planilha
        for campo in planilha['b']:
            iterador += 1
            if campo.value == 'Pago':
                campo.fill = greenFill
            elif campo.value == 'Pendente':
                campo.fill = yellowFill
            elif iterador > 1:
                campo.fill = redFill

class Endereco:
        def recebe_endereco(self, campo, parte):
            try:
                if parte == 1:
                    end = campo['shipping_address']['address1'].replace(' ', '').replace('-', '').replace(',',
                                                                                                          '').replace(
                        '.', '').replace('(', '').replace(')', '').replace('/', '')
                else:
                    end = campo['shipping_address']['address2'].replace(' ', '').replace('-', '').replace(',',
                                                                                                          '').replace(
                        '.', '').replace('(', '').replace(')', '').replace('/', '')
                return end
            except:
                pass

        def get_endereco_de_outra_forma(self, campo, parte):
            try:
                if parte == 1:
                    end = campo['billing_address']['address1'].replace(' ', '').replace('-', '').replace(',',
                                                                                                         '').replace(
                        '.', '').replace('(', '').replace(')', '').replace('/', '')
                else:
                    end = campo['billing_address']['address2'].replace(' ', '').replace('-', '').replace(',',
                                                                                                         '').replace(
                        '.', '').replace('(', '').replace(')', '').replace('/', '')
                return end
            except:
                pass

class Pedido:
        def get_numero_pedido(self, campo):
            try:
                return campo['order_number']
            except:
                pass

        def get_forma_pagamento(self, campo):
            try:
                if 'paypal' in campo['gateway']:
                    return 'Paypal'
                else:
                    return 'PagSeguro'
            except:
                pass

        def get_status_pagamento(self, campo):
            try:
                if campo['financial_status'] == 'paid':
                    return 'Pago'
                else:
                    return 'Pendente'
            except:
                pass

        def get_andamento_pedido(self, campo):
            try:
                return campo['fulfillment_status']
            except:
                pass

        def get_forma_envio(self, campo):
            try:
                for campo in campo["shipping_lines"]:
                    return campo['title']
            except:
                pass

        def get_email_cliente(self, campo):
            try:
                return campo['customer']['email']
            except:
                pass

        def get_nome_cliente(self, campo):
            try:
                return campo['billing_address']['name']
            except:
                pass

        def get_quantidade_itens(self, campo):
            try:
                items = 0
                for campo in campo["line_items"]:
                    items = items + campo['quantity']
                return items
            except:
                pass
